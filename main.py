# Python program to read an excel file

# import openpyxl module
import openpyxl
import csv
# Give the location of the file
path = "csv/retour.xlsx"

# To open the workbook
# workbook object is created
wb_obj = openpyxl.load_workbook(path)

# Get workbook active sheet object
# from the active attribute
ws = wb_obj.active


# Print value of cell object
# using the value attribute

tab = []
i=1
with open('csv/planifRit.csv', ) as csvfile:
    spamreader = csv.reader(csvfile, delimiter=';')
    for index,ligne in enumerate(spamreader):

        if ws.cell(row=i, column=9).value == ligne[0]:

            print(ws.cell(row=i, column=16).value)
            ws.cell(row=i, column=16).value = ligne[1] +" "+ligne[2]
        i = i +1
    wb_obj.save("RetourOrange.xlsx")
    print("C'est un succes")
